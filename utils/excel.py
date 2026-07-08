from pathlib import Path
from shutil import copy2
from openpyxl import load_workbook

# ============================================================
# PASTAS DO PROJETO
# ============================================================

PASTA_MODELOS = Path("modelos")
PASTA_PACIENTES = Path("pacientes")

MODELO = PASTA_MODELOS / "modelo_paciente.xlsx"

PASTA_MODELOS.mkdir(exist_ok=True)
PASTA_PACIENTES.mkdir(exist_ok=True)


# ============================================================
# LISTAR PACIENTES
# ============================================================

def listar_pacientes():

    pacientes = []

    for arquivo in sorted(PASTA_PACIENTES.glob("*.xlsx")):

        pacientes.append(arquivo.stem)

    return pacientes


# ============================================================
# CAMINHO DO PACIENTE
# ============================================================

def caminho_paciente(nome):

    return PASTA_PACIENTES / f"{nome}.xlsx"


# ============================================================
# PACIENTE EXISTE
# ============================================================

def paciente_existe(nome):

    return caminho_paciente(nome).exists()


# ============================================================
# COPIAR MODELO
# ============================================================

def copiar_modelo(nome):

    destino = caminho_paciente(nome)

    copy2(MODELO, destino)

    return destino


# ============================================================
# CRIAR PACIENTE
# ============================================================

def criar_paciente(
        nome,
        cpf="",
        nascimento="",
        sexo="",
        telefone="",
        responsavel="",
        escola="",
        diagnostico=""
):

    if nome.strip() == "":

        return False

    if paciente_existe(nome):

        return False

    destino = copiar_modelo(nome)

    wb = load_workbook(destino)

    ws = wb["Cadastro"]

    ws["B4"] = nome
    ws["B5"] = cpf
    ws["B6"] = nascimento
    ws["B7"] = sexo
    ws["B8"] = telefone
    ws["B9"] = responsavel
    ws["B10"] = escola
    ws["B11"] = diagnostico

    wb.save(destino)

    wb.close()

    return True


# ============================================================
# ABRIR WORKBOOK
# ============================================================

def abrir_paciente(nome):

    arquivo = caminho_paciente(nome)

    if not arquivo.exists():

        return None

    return load_workbook(arquivo)


# ============================================================
# LER CADASTRO
# ============================================================

def ler_cadastro(nome):

    wb = abrir_paciente(nome)

    if wb is None:

        return None

    ws = wb["Cadastro"]

    dados = {

        "Nome": ws["B4"].value,
        "CPF": ws["B5"].value,
        "Nascimento": ws["B6"].value,
        "Sexo": ws["B7"].value,
        "Telefone": ws["B8"].value,
        "Responsável": ws["B9"].value,
        "Escola": ws["B10"].value,
        "Diagnóstico": ws["B11"].value

    }

    wb.close()

    return dados# ============================================================
# SALVAR CADASTRO
# ============================================================

def salvar_cadastro(nome, dados):

    wb = abrir_paciente(nome)

    if wb is None:
        return False

    ws = wb["Cadastro"]

    ws["B4"] = dados.get("Nome", "")
    ws["B5"] = dados.get("CPF", "")
    ws["B6"] = dados.get("Nascimento", "")
    ws["B7"] = dados.get("Sexo", "")
    ws["B8"] = dados.get("Telefone", "")
    ws["B9"] = dados.get("Responsável", "")
    ws["B10"] = dados.get("Escola", "")
    ws["B11"] = dados.get("Diagnóstico", "")

    wb.save(caminho_paciente(nome))
    wb.close()

    return True


# ============================================================
# EXCLUIR PACIENTE
# ============================================================

def excluir_paciente(nome):

    arquivo = caminho_paciente(nome)

    if arquivo.exists():

        arquivo.unlink()

        return True

    return False


# ============================================================
# LER AVALIAÇÕES
# ============================================================

def ler_avaliacao(nome):

    wb = abrir_paciente(nome)

    if wb is None:
        return []

    if "Avaliacoes" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["Avaliacoes"]

    lista = []

    linha = 2

    while True:

        dominio = ws[f"A{linha}"].value

        if dominio is None:
            break

        lista.append({

            "Domínio": dominio,
            "Avaliação 1": ws[f"B{linha}"].value,
            "Avaliação 2": ws[f"C{linha}"].value,
            "Avaliação 3": ws[f"D{linha}"].value,
            "Avaliação 4": ws[f"E{linha}"].value

        })

        linha += 1

    wb.close()

    return lista


# ============================================================
# SALVAR AVALIAÇÕES
# ============================================================

def salvar_avaliacao(nome, lista):

    wb = abrir_paciente(nome)

    if wb is None:
        return False

    if "Avaliacoes" not in wb.sheetnames:
        wb.close()
        return False

    ws = wb["Avaliacoes"]

    linha = 2

    for item in lista:

        ws[f"B{linha}"] = item["Avaliação 1"]
        ws[f"C{linha}"] = item["Avaliação 2"]
        ws[f"D{linha}"] = item["Avaliação 3"]
        ws[f"E{linha}"] = item["Avaliação 4"]

        linha += 1

    wb.save(caminho_paciente(nome))

    wb.close()

    return True


# ============================================================
# CONTADORES
# ============================================================

def total_pacientes():

    return len(listar_pacientes())


# ============================================================
# ÚLTIMO PACIENTE
# ============================================================

def ultimo_paciente():

    lista = listar_pacientes()

    if len(lista) == 0:
        return ""

    return lista[-1]


# ============================================================
# DASHBOARD
# ============================================================

def resumo():

    return {

        "total_pacientes": total_pacientes(),

        "ultimo_paciente": ultimo_paciente()

    }
from openpyxl import load_workbook
from pathlib import Path

PASTA_PACIENTES = Path("pacientes")

def abrir_planilha(nome):
    arquivo = PASTA_PACIENTES / f"{nome}.xlsx"
    return load_workbook(arquivo)
# =============================================
# LER UMA CÉLULA
# =============================================

def ler(nome, aba, celula):

    wb = abrir_planilha(nome)

    ws = wb[aba]

    valor = ws[celula].value

    wb.close()

    return valor


# =============================================
# ESCREVER UMA CÉLULA
# =============================================

def escrever(nome, aba, celula, valor):

    wb = abrir_planilha(nome)

    ws = wb[aba]

    ws[celula] = valor

    wb.save(caminho_paciente(nome))

    wb.close()
    # =============================================
# CALCULAR SCORING MOTAS
# =============================================

def calcular_scoring(nome):

    itens = [
        "B2","B3","B4","B5","B6",
        "B7","B8","B9","B10","B11","B12"
    ]

    total = 0

    for celula in itens:

        valor = ler(nome, "MOTAS", celula)

        if valor is None or valor == "":
            valor = 0

        valor = float(valor)

        escrever(nome, "MOTAS_SCORING", celula, valor)

        total += valor

    escrever(nome, "MOTAS_SCORING", "B13", total)

    return total